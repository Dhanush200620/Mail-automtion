# create_yearly_tracker.py
# Generates a one-year leave calculation Excel tracker (Jul 2026 – Jun 2027)
# with professional formatting, formulas, and all team members.
#
# Run this script directly:  python create_yearly_tracker.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
import os
from datetime import date

# ============================================================
# Configuration
# ============================================================

# Year range: Jul 2026 to Jun 2029 (3 years)
START_YEAR = 2026
START_MONTH = 7
END_YEAR = 2029
END_MONTH = 6

OUTPUT_FILE = "leave_tracker_2026_2029.xlsx"
# Also update the test file used by the project
TEST_FILE = "test_leave_tracker.xlsx"

# Full team list — updated based on provided image
TEAM = [
    ("MDM-BAU", "Kavivanan"),
    ("MDM-BAU", "Leando Iruthayaraj"),
    ("MDM-BAU", "Vibra Narayanan"),
    ("MDM-BAU", "Aravind"),
    ("MDM-BAU", "Jaya Shree"),
    ("MDM - GROW", "Vignesh Pugalendhi"),
    ("MDM - GROW", "Sujitha"),
    ("MDM - GROW", "Balaji Loganathan"),
    ("MDM - GROW", "Ebron Stalin"),
    ("MDM - GROW", "Rahul Kumar"),
    ("MDM - GROW", "Harivarshan"),
    ("MDM - GROW", "BharaniDharan"),
    ("MDM - GROW", "Sandhiya"),
    ("MDM - GROW", "Jose"),
    ("CRM - Next Gen", "Dravid"),
    ("CRM - Next Gen", "Soma Sai Pavan"),
    ("CRM - Next Gen", "Rathimeena"),
    ("CRM - Next Gen", "Pavithra"),
    ("CRM - Next Gen", "Jefflina"),
    ("CRM - Next Gen", "Manoj"),
    ("CDS", "Arshad"),
    ("CDS", "Dushyanth"),
    ("PowerBI / Tableau / Alteryx / Devops Admin", "Hari Vembeina"),
    ("PowerBI / Tableau / Alteryx / Devops Admin", "JP"),
    ("PowerBI / Tableau / Alteryx / Devops Admin", "Jelsi"),
    ("PowerBI / Tableau / Alteryx / Devops Admin", "Mahesh Reddy"),
    ("PowerBI / Tableau / Alteryx / Devops Admin", "Sandeep Reddy"),
    ("", "Alafia Yusuf Fidvi"),
    ("", "Arun R"),
    ("", "Swaminathan"),
    ("", "Vasanth"),
]

# Indian public holidays (2026-2027) — adjust as needed
INDIAN_HOLIDAYS = {
    date(2026, 8, 15): "Independence Day",
    date(2026, 10, 2): "Gandhi Jayanti",
    date(2026, 10, 20): "Diwali",
    date(2026, 10, 21): "Diwali (Day 2)",
    date(2026, 11, 4): "Diwali Amavasya",
    date(2026, 12, 25): "Christmas",
    date(2027, 1, 1): "New Year",
    date(2027, 1, 14): "Pongal",
    date(2027, 1, 15): "Pongal (Day 2)",
    date(2027, 1, 26): "Republic Day",
    date(2027, 3, 29): "Holi",
    date(2027, 4, 2): "Good Friday",
    date(2027, 4, 14): "Tamil New Year",
    date(2027, 5, 1): "May Day",
}

# Total leave entitlements per year
LEAVE_ENTITLEMENTS = {
    "CL": 12,  # Casual Leave
    "SL": 12,  # Sick Leave
    "PL": 15,  # Privilege Leave (Earned Leave)
    "EL": 0,   # Emergency Leave (as needed, no fixed quota)
}

# ============================================================
# Styling
# ============================================================

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SUB_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SUB_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
WEEKEND_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
HOLIDAY_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TOTAL_HEADER_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
TOTAL_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NAME_FONT = Font(name="Calibri", bold=True, size=10)
CELL_FONT = Font(name="Calibri", size=9)
STREAM_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


def generate_months():
    """Generate list of (year, month) tuples for the tracker period."""
    months = []
    y, m = START_YEAR, START_MONTH
    while (y, m) <= (END_YEAR, END_MONTH):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def create_month_sheet(wb, year, month_num):
    """Create one month sheet with full formatting and formulas."""
    month_abbr = calendar.month_abbr[month_num]
    sheet_name = f"{month_abbr} {year}"
    ws = wb.create_sheet(title=sheet_name)

    days_in_month = calendar.monthrange(year, month_num)[1]

    # ---- Row 1: Headers (Stream, Name, Day names, Totals) ----
    ws.cell(row=1, column=1, value="Stream")
    ws.cell(row=1, column=2, value="Team Member Name")
    for c in [1, 2]:
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    for day in range(1, days_in_month + 1):
        col = day + 2
        d = date(year, month_num, day)
        cell = ws.cell(row=1, column=col, value=d.strftime("%a"))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Total columns
    total_start = days_in_month + 3
    total_headers = [
        "Total CL", "Total SL", "Total PL", "Total EL", "Total HD",
        "Total WFH", "Total Week Off", "Total Comp Off", "Total Leave Days"
    ]
    for idx, header in enumerate(total_headers):
        cell = ws.cell(row=1, column=total_start + idx, value=header)
        cell.font = TOTAL_HEADER_FONT
        cell.fill = TOTAL_HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # ---- Row 2: Day numbers ----
    ws.cell(row=2, column=1, value="")
    ws.cell(row=2, column=2, value="")
    for c in [1, 2]:
        cell = ws.cell(row=2, column=c)
        cell.fill = SUB_HEADER_FILL
        cell.border = THIN_BORDER

    for day in range(1, days_in_month + 1):
        col = day + 2
        d = date(year, month_num, day)
        cell = ws.cell(row=2, column=col, value=day)
        cell.font = SUB_HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

        # Color weekends and holidays in header row too
        if d.weekday() in [5, 6]:
            cell.fill = WEEKEND_FILL
            cell.font = Font(name="Calibri", bold=True, color="4472C4", size=9)
        elif d in INDIAN_HOLIDAYS:
            cell.fill = HOLIDAY_FILL
            cell.font = Font(name="Calibri", bold=True, color="C00000", size=9)
        else:
            cell.fill = SUB_HEADER_FILL

    # Style total columns row 2
    for idx in range(len(total_headers)):
        cell = ws.cell(row=2, column=total_start + idx, value="")
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    # ---- Row 3+: Team members ----
    for i, (stream, name) in enumerate(TEAM):
        row = i + 3

        # Stream column
        stream_cell = ws.cell(row=row, column=1, value=stream)
        stream_cell.font = CELL_FONT
        stream_cell.fill = STREAM_FILL
        stream_cell.alignment = LEFT_ALIGN
        stream_cell.border = THIN_BORDER

        # Name column
        name_cell = ws.cell(row=row, column=2, value=name)
        name_cell.font = NAME_FONT
        name_cell.alignment = LEFT_ALIGN
        name_cell.border = THIN_BORDER

        # Day cells
        for day in range(1, days_in_month + 1):
            d = date(year, month_num, day)
            col = day + 2
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

            if d.weekday() in [5, 6]:
                cell.value = "WO"
                cell.fill = WEEKEND_FILL
                cell.font = Font(name="Calibri", size=9, color="4472C4")
            elif d in INDIAN_HOLIDAYS:
                cell.value = "OH"
                cell.fill = HOLIDAY_FILL
                cell.font = Font(name="Calibri", size=9, color="C00000")

        # ---- Totals formulas ----
        last_data_col = get_column_letter(days_in_month + 2)
        formulas = [
            f'=COUNTIF(C{row}:{last_data_col}{row}, "CL")',   # Total CL
            f'=COUNTIF(C{row}:{last_data_col}{row}, "SL")',   # Total SL
            f'=COUNTIF(C{row}:{last_data_col}{row}, "PL")',   # Total PL
            f'=COUNTIF(C{row}:{last_data_col}{row}, "EL")',   # Total EL
            f'=COUNTIF(C{row}:{last_data_col}{row}, "HD")',   # Total HD
            f'=COUNTIF(C{row}:{last_data_col}{row}, "WFH")',  # Total WFH
            f'=COUNTIF(C{row}:{last_data_col}{row}, "WO")',   # Total Week Off
            f'=COUNTIF(C{row}:{last_data_col}{row}, "CO")',   # Total Comp Off
        ]
        # Total Leave Days = CL + SL + PL + EL + (HD * 0.5)
        cl_col = get_column_letter(total_start)
        el_col = get_column_letter(total_start + 3)
        hd_col = get_column_letter(total_start + 4)
        formulas.append(
            f'=SUM({cl_col}{row}:{el_col}{row}) + ({hd_col}{row}*0.5)'
        )

        for idx, formula in enumerate(formulas):
            cell = ws.cell(row=row, column=total_start + idx, value=formula)
            cell.font = Font(name="Calibri", bold=True, size=9)
            cell.fill = SUMMARY_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    for day in range(1, days_in_month + 1):
        ws.column_dimensions[get_column_letter(day + 2)].width = 5
    for idx in range(len(total_headers)):
        ws.column_dimensions[get_column_letter(total_start + idx)].width = 14

    # Freeze panes: freeze columns A & B and rows 1 & 2
    ws.freeze_panes = "C3"

    return ws


def create_summary_sheet(wb, months):
    """Create a yearly summary sheet that aggregates totals from all monthly sheets."""
    ws = wb.create_sheet(title="Yearly Summary", index=0)

    # ---- Title ----
    title_cell = ws.cell(row=1, column=1, value="LEAVE TRACKER — YEARLY SUMMARY")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:P1")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    period_cell = ws.cell(row=2, column=1, value=f"Period: {calendar.month_abbr[START_MONTH]} {START_YEAR} – {calendar.month_abbr[END_MONTH]} {END_YEAR}")
    period_cell.font = Font(name="Calibri", size=11, italic=True, color="4472C4")
    ws.merge_cells("A2:P2")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # ---- Headers (Row 4) ----
    headers = [
        "S.No", "Stream", "Team Member Name",
        "CL Taken", "SL Taken", "PL Taken", "EL Taken", "HD Taken",
        "Total Leave Days", "WFH Days", "Week Offs", "Comp Offs",
        "CL Balance", "SL Balance", "PL Balance"
    ]
    for c_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # ---- Data rows (Row 5+) ----
    for i, (stream, name) in enumerate(TEAM):
        row = i + 5

        ws.cell(row=row, column=1, value=i + 1).font = CELL_FONT
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        stream_cell = ws.cell(row=row, column=2, value=stream)
        stream_cell.font = CELL_FONT
        stream_cell.fill = STREAM_FILL
        stream_cell.alignment = LEFT_ALIGN
        stream_cell.border = THIN_BORDER

        name_cell = ws.cell(row=row, column=3, value=name)
        name_cell.font = NAME_FONT
        name_cell.alignment = LEFT_ALIGN
        name_cell.border = THIN_BORDER

        # Build sum formulas across all 12 monthly sheets
        # Person is always in the same row across sheets (row = i + 3)
        person_row_in_sheets = i + 3

        # For each monthly sheet, the total columns start at (days_in_month + 3)
        # We need to reference the correct column in each sheet.
        # Since days_in_month varies, we'll build individual cell references.

        leave_types = ["CL", "SL", "PL", "EL", "HD", "WFH", "WO", "CO"]
        total_col_offsets = {
            "CL": 0, "SL": 1, "PL": 2, "EL": 3,
            "HD": 4, "WFH": 5, "WO": 6, "CO": 7, "TOTAL": 8,
        }

        def build_sum_formula(offset):
            """Sum a specific total column across all 12 monthly sheets."""
            parts = []
            for (y, m) in months:
                days = calendar.monthrange(y, m)[1]
                t_col = get_column_letter(days + 3 + offset)
                sheet = f"{calendar.month_abbr[m]} {y}"
                parts.append(f"'{sheet}'!{t_col}{person_row_in_sheets}")
            return "=" + "+".join(parts)

        # Columns D-H: CL, SL, PL, EL, HD Taken
        for col_offset, dest_col in enumerate([4, 5, 6, 7, 8]):
            cell = ws.cell(row=row, column=dest_col)
            cell.value = build_sum_formula(col_offset)
            cell.font = Font(name="Calibri", bold=True, size=10)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Column I: Total Leave Days
        cell = ws.cell(row=row, column=9)
        cell.value = build_sum_formula(total_col_offsets["TOTAL"])
        cell.font = Font(name="Calibri", bold=True, size=10, color="C00000")
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

        # Column J: WFH Days
        cell = ws.cell(row=row, column=10)
        cell.value = build_sum_formula(total_col_offsets["WFH"])
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

        # Column K: Week Offs
        cell = ws.cell(row=row, column=11)
        cell.value = build_sum_formula(total_col_offsets["WO"])
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

        # Column L: Comp Offs
        cell = ws.cell(row=row, column=12)
        cell.value = build_sum_formula(total_col_offsets["CO"])
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

        # Columns M-O: Leave Balances (Entitlement - Taken)
        for balance_col, leave_type, taken_col in [
            (13, "CL", "D"), (14, "SL", "E"), (15, "PL", "F")
        ]:
            entitlement = LEAVE_ENTITLEMENTS[leave_type]
            cell = ws.cell(row=row, column=balance_col)
            cell.value = f"={entitlement}-{taken_col}{row}"
            cell.font = Font(name="Calibri", bold=True, size=10)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            # Conditional-style: green fill for positive balance
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # ---- Entitlement Reference Row ----
    ref_row = len(TEAM) + 6
    ws.cell(row=ref_row, column=1, value="").border = THIN_BORDER
    label = ws.cell(row=ref_row, column=2, value="Leave Entitlements (per year):")
    label.font = Font(name="Calibri", bold=True, italic=True, size=10, color="1F4E79")
    ws.merge_cells(start_row=ref_row, start_column=2, end_row=ref_row, end_column=3)

    for ref_col, (code, quota) in enumerate(LEAVE_ENTITLEMENTS.items(), start=4):
        cell = ws.cell(row=ref_row, column=ref_col, value=f"{code}: {quota}")
        cell.font = Font(name="Calibri", italic=True, size=10, color="548235")
        cell.alignment = CENTER_ALIGN

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    for c in range(4, 16):
        ws.column_dimensions[get_column_letter(c)].width = 15

    ws.freeze_panes = "D5"

    return ws


def create_leave_tracker():
    """Main function: creates the full one-year leave tracker workbook."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, OUTPUT_FILE)
    test_path = os.path.join(script_dir, TEST_FILE)

    print("=" * 60)
    print("  Creating One-Year Leave Tracker")
    print(f"  Period: {calendar.month_abbr[START_MONTH]} {START_YEAR} – {calendar.month_abbr[END_MONTH]} {END_YEAR}")
    print(f"  Team members: {len(TEAM)}")
    print("=" * 60)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Generate all months
    months = generate_months()

    # Create monthly sheets
    for year, month_num in months:
        sheet_name = f"{calendar.month_abbr[month_num]} {year}"
        print(f"  Creating sheet: {sheet_name}...")
        create_month_sheet(wb, year, month_num)

    # Create yearly summary (inserted at position 0)
    print("  Creating Yearly Summary sheet...")
    create_summary_sheet(wb, months)

    # Save both files
    wb.save(output_path)
    wb.save(test_path)
    print(f"\n{'=' * 60}")
    print(f"  Done! Saved to:")
    print(f"    - {output_path}")
    print(f"    - {test_path}")
    print(f"  Sheets created: {len(wb.sheetnames)}")
    print(f"  Sheet names: {', '.join(wb.sheetnames)}")
    print(f"{'=' * 60}")

    return output_path


if __name__ == "__main__":
    create_leave_tracker()
