# update_excel.py
# This script handles reading and writing leave codes into the Excel tracker.
#
# The Excel workbook has one sheet per month (like "Jun 2026", "Jul 2026").
# Column A = Stream (team name), Column B = Team Member Name
# Row 1 = day names (Mon, Tue, ...), Row 2 = day numbers (1, 2, 3, ...)
# Row 3 onward = one row per person
# Last 4 columns of each sheet are totals — we don't touch those.
#
# I'm using openpyxl to read/write Excel and rapidfuzz for fuzzy name matching.

import openpyxl
import json
import os
import calendar
from datetime import datetime, date, timedelta
from rapidfuzz import fuzz
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# --- load config ---
# keeping all settings in config.json so I can tweak stuff without changing code
script_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(script_dir, "config.json")

with open(config_path) as f:
    config = json.load(f)

# codes we should never overwrite (WFH = work from home, WO = week off, etc.)
PROTECTED_CODES = config.get("protected_codes", ["WFH", "WS", "WO", "OH", "HL"])
LEAVE_CODES = ["SL", "CL", "PL", "EL", "HD"]
FUZZY_THRESHOLD = config.get("fuzzy_threshold", 75)

# Google Sheets Configuration
USE_GOOGLE_SHEETS = config.get("use_google_sheets", False)
GOOGLE_SHEET_NAME = config.get("google_sheet_name", "Leave Tracker")
GOOGLE_CREDENTIALS_FILE = config.get("google_credentials_file", "google_credentials.json")

_gspread_client = None

def get_gspread_client():
    global _gspread_client
    if _gspread_client is None:
        creds_file = GOOGLE_CREDENTIALS_FILE
        if not os.path.isabs(creds_file):
            creds_file = os.path.join(script_dir, creds_file)
            
        if not os.path.exists(creds_file):
            raise FileNotFoundError(
                f"Google credentials file not found at {creds_file}. "
                "Please configure 'google_credentials_file' in config.json "
                "and ensure the JSON credentials file exists."
            )
            
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(creds_file, scope)
        _gspread_client = gspread.authorize(creds)
    return _gspread_client


# ============================================================
# Helper functions
# ============================================================

def get_sheet_name(d):
    # turns a date object into the sheet name format the workbook uses
    # e.g. date(2026, 6, 15) -> "Jun 2026"
    return d.strftime("%b %Y")


def find_person_row(ws, name):
    # look through column B (team member names) and find the closest match
    # using fuzzy matching because email names don't always match exactly
    # e.g. "Dhanush K" in email vs "Dhanush Kumar" in the sheet
    
    best_row = None
    best_score = 0
    best_name = None
    
    for row in range(3, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is None:
            continue
        
        cell_name = str(cell_val).strip()
        
        # WRatio tries multiple matching strategies and picks the best one
        # it handles partial matches, word reordering, etc.
        # TODO: maybe I should also try token_sort_ratio? need to test more
        score = fuzz.WRatio(name.lower().strip(), cell_name.lower())
        
        if score > best_score:
            best_score = score
            best_row = row
            best_name = cell_name
    
    if best_score >= FUZZY_THRESHOLD:
        print(f"  Matched '{name}' -> '{best_name}' (score: {best_score})")
        return best_row, best_name, best_score
    else:
        print(f"  WARNING: no good match for '{name}' (best was '{best_name}' at {best_score})")
        return None, best_name, best_score


def find_date_column(ws, target_date):
    # Row 2 has the day numbers (1, 2, 3, ...)
    # we just match against the day-of-month since the sheet name tells us
    # which month we're in already
    
    day = target_date.day
    
    for col in range(3, ws.max_column + 1):
        cell_val = ws.cell(row=2, column=col).value
        if cell_val is None:
            continue
        
        # the cell could be an int, float, string, or even a datetime object
        # gotta handle all the weird cases openpyxl can throw at you
        if isinstance(cell_val, (int, float)):
            if int(cell_val) == day:
                return col
        elif isinstance(cell_val, str):
            # sometimes dates get saved as strings... annoying
            try:
                if int(cell_val.strip()) == day:
                    return col
            except ValueError:
                pass
        elif isinstance(cell_val, datetime):
            if cell_val.day == day:
                return col
    
    print(f"  ERROR: couldn't find column for day {day}")
    return None


def find_person_row_from_values(values, name):
    # look through column B (team member names) and find the closest match
    best_row = None
    best_score = 0
    best_name = None
    
    for i in range(2, len(values)):
        row_vals = values[i]
        if len(row_vals) < 2:
            continue
        cell_val = row_vals[1]
        if cell_val is None:
            continue
        
        cell_name = str(cell_val).strip()
        score = fuzz.WRatio(name.lower().strip(), cell_name.lower())
        
        if score > best_score:
            best_score = score
            best_row = i + 1
            best_name = cell_name
            
    if best_score >= FUZZY_THRESHOLD:
        print(f"  Matched '{name}' -> '{best_name}' (score: {best_score})")
        return best_row, best_name, best_score
    else:
        print(f"  WARNING: no good match for '{name}' (best was '{best_name}' at {best_score})")
        return None, best_name, best_score


def find_date_column_from_values(values, target_date):
    day = target_date.day
    if len(values) < 2:
        return None
    row2 = values[1]
    
    for c_idx in range(2, len(row2)):
        cell_val = row2[c_idx]
        if cell_val is None or cell_val == "":
            continue
            
        try:
            if int(float(str(cell_val).strip())) == day:
                return c_idx + 1
        except ValueError:
            pass
            
    print(f"  ERROR: couldn't find column for day {day}")
    return None


# ============================================================
# Main write function
# ============================================================

def write_leave(excel_path, person_name, leave_date, leave_code, is_correction=False):
    # writes a leave code into the correct cell
    # returns (True/False, message) so the caller knows what happened
    
    sheet_name = get_sheet_name(leave_date)
    
    if USE_GOOGLE_SHEETS:
        try:
            client = get_gspread_client()
            sh = client.open(GOOGLE_SHEET_NAME)
        except FileNotFoundError as e:
            print(f"  ERROR: Google credentials file not found: {e}")
            return False, str(e)
        except Exception as e:
            print(f"  ERROR: failed to connect to Google Sheets '{GOOGLE_SHEET_NAME}': {e}")
            return False, f"Failed to connect to Google Sheets: {e}"
            
        try:
            ws = sh.worksheet(sheet_name)
        except Exception:
            print(f"  ERROR: sheet '{sheet_name}' doesn't exist in the Google Sheet")
            return False, f"Sheet '{sheet_name}' not found"
            
        try:
            values = ws.get_all_values()
        except Exception as e:
            print(f"  ERROR: failed to read values from sheet '{sheet_name}': {e}")
            return False, f"Failed to read sheet: {e}"
            
        row, matched_name, score = find_person_row_from_values(values, person_name)
        if row is None:
            return False, f"No match for '{person_name}' (best: '{matched_name}' at {score}%)"
            
        col = find_date_column_from_values(values, leave_date)
        if col is None:
            return False, f"Date {leave_date.strftime('%d %b')} not found in sheet"
            
        current_val = None
        if len(values) >= row and len(values[row-1]) >= col:
            current_val = values[row-1][col-1]
            
        if current_val:
            current_str = str(current_val).strip().upper()
            
            if current_str in [p.upper() for p in PROTECTED_CODES]:
                print(f"  SKIP: {matched_name} on {leave_date.strftime('%d %b')} — cell has '{current_val}' (protected)")
                return False, f"Protected code '{current_val}' in cell"
                
            if current_str in [c.upper() for c in LEAVE_CODES]:
                if not is_correction:
                    print(f"  SKIP: {matched_name} on {leave_date.strftime('%d %b')} — already has '{current_val}'")
                    return False, f"Already has leave code '{current_val}'"
                else:
                    print(f"  CORRECTION: overwriting '{current_val}' with '{leave_code}' for {matched_name}")
                    
        try:
            ws.update_cell(row, col, leave_code)
        except Exception as e:
            print(f"  ERROR: failed to write to Google Sheets: {e}")
            return False, f"Failed to update Google Sheets: {e}"
            
        print(f"  WROTE (Google Sheets): '{leave_code}' for {matched_name} on {leave_date.strftime('%d %b %Y')}")
        return True, f"Wrote {leave_code} for {matched_name}"
        
    else:
        # local Excel openpyxl workflow
        if not os.path.exists(excel_path):
            print(f"  ERROR: can't find the Excel file at {excel_path}")
            return False, "Excel file not found"
        
        wb = openpyxl.load_workbook(excel_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"  ERROR: sheet '{sheet_name}' doesn't exist in the workbook")
            wb.close()
            return False, f"Sheet '{sheet_name}' not found"
        
        ws = wb[sheet_name]
        
        row, matched_name, score = find_person_row(ws, person_name)
        if row is None:
            wb.close()
            return False, f"No match for '{person_name}' (best: '{matched_name}' at {score}%)"
        
        col = find_date_column(ws, leave_date)
        if col is None:
            wb.close()
            return False, f"Date {leave_date.strftime('%d %b')} not found in sheet"
        
        current_val = ws.cell(row=row, column=col).value
        
        if current_val:
            current_str = str(current_val).strip().upper()
            
            if current_str in [p.upper() for p in PROTECTED_CODES]:
                print(f"  SKIP: {matched_name} on {leave_date.strftime('%d %b')} — cell has '{current_val}' (protected)")
                wb.close()
                return False, f"Protected code '{current_val}' in cell"
            
            if current_str in [c.upper() for c in LEAVE_CODES]:
                if not is_correction:
                    print(f"  SKIP: {matched_name} on {leave_date.strftime('%d %b')} — already has '{current_val}'")
                    wb.close()
                    return False, f"Already has leave code '{current_val}'"
                else:
                    print(f"  CORRECTION: overwriting '{current_val}' with '{leave_code}' for {matched_name}")
        
        ws.cell(row=row, column=col).value = leave_code
        wb.save(excel_path)
        wb.close()
        
        print(f"  WROTE: '{leave_code}' for {matched_name} on {leave_date.strftime('%d %b %Y')}")
        return True, f"Wrote {leave_code} for {matched_name}"


def write_leave_multi_day(excel_path, person_name, start_date, end_date, leave_code, is_correction=False):
    # handles leave that spans multiple days (and possibly multiple months)
    # just loops through each day and calls write_leave()
    # it's simple but it works — each day gets processed individually
    
    results = []
    current = start_date
    
    while current <= end_date:
        print(f"\n  Day: {current.strftime('%a %d %b %Y')}...")
        success, msg = write_leave(excel_path, person_name, current, leave_code, is_correction)
        results.append({
            "date": current,
            "success": success,
            "message": msg
        })
        current += timedelta(days=1)
    
    return results


# ============================================================
# Test Excel generator
# ============================================================
# This creates a sample Excel file that mimics the real format
# so I can test everything without needing the actual workbook

def create_test_excel(path):
    print("Creating test Excel file with the real MDM BAU format...")
    
    wb = openpyxl.Workbook()
    # openpyxl creates a default "Sheet" — remove it since we'll add our own
    wb.remove(wb.active)
    
    # sample team members (made up names for testing)
    team = [
        ("MDM-BAU", "Kavivanan"),
        ("MDM-BAU", "Leando Iruthayaraj"),
        ("MDM-BAU", "Vibra Narayanan"),
        ("MDM-BAU", "Aravind"),
        ("MDM-BAU", "Jaya Shree"),
        ("MDM - GROW", "Vignesh Pugalendhi"),
        ("MDM - GROW", "Sujitha"),
        ("MDM - GROW", "Balaji Loganathan"),
        ("MDM - GROW", "Ebron Stalin"),
        ("MDM - GROW", "Rahul Kumar"),
        ("MDM - GROW", "Harivarshan"),
        ("MDM - GROW", "BharaniDharan"),
        ("MDM - GROW", "Sandhiya"),
        ("MDM - GROW", "Jose"),
        ("CRM - Next Gen", "Dravid"),
        ("CRM - Next Gen", "Soma Sai Pavan"),
        ("CRM - Next Gen", "Rathimeena"),
        ("CRM - Next Gen", "Pavithra"),
        ("CRM - Next Gen", "Jefflina"),
        ("CRM - Next Gen", "Manoj"),
        ("CDS", "Arshad"),
        ("CDS", "Dushyanth"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "Hari Vembeina"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "JP"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "Jelsi"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "Mahesh Reddy"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "Sandeep Reddy"),
        ("", "Alafia Yusuf Fidvi"),
        ("", "Arun R"),
        ("", "Swaminathan"),
        ("", "Vasanth"),
    ]
    
    # create sheets for Jun and Jul 2026 so we can test multi-month stuff
    for month_num in [6, 7]:
        year = 2026
        month_abbr = calendar.month_abbr[month_num]  # "Jun", "Jul"
        sheet_name = f"{month_abbr} {year}"
        
        ws = wb.create_sheet(title=sheet_name)
        
        days_in_month = calendar.monthrange(year, month_num)[1]
        
        # --- Row 1: "Stream", "Team Member Name", then day-of-week abbreviations ---
        ws.cell(row=1, column=1).value = "Stream"
        ws.cell(row=1, column=2).value = "Team Member Name"
        
        for day in range(1, days_in_month + 1):
            col = day + 2  # +2 because col A=stream, col B=name
            d = date(year, month_num, day)
            ws.cell(row=1, column=col).value = d.strftime("%a")  # Mon, Tue, etc.
        
        # --- Row 2: blank, blank, then day numbers (1, 2, 3, ...) ---
        for day in range(1, days_in_month + 1):
            col = day + 2
            ws.cell(row=2, column=col).value = day
        
        # --- Totals headers in the last 9 columns ---
        total_start = days_in_month + 3
        headers = [
            "Total CL", "Total SL", "Total PL", "Total EL", "Total HD",
            "Total WFH", "Total Week Off", "Total Comp Off", "Total Leave Days"
        ]
        for idx, header in enumerate(headers):
            ws.cell(row=1, column=total_start + idx).value = header
        
        # --- Row 3+: fill in team members ---
        for i, (stream, name) in enumerate(team):
            row = i + 3  # start from row 3
            ws.cell(row=row, column=1).value = stream
            ws.cell(row=row, column=2).value = name
            
            # mark Saturdays and Sundays as WO (week off)
            for day in range(1, days_in_month + 1):
                d = date(year, month_num, day)
                col = day + 2
                if d.weekday() in [5, 6]:  # 5=Saturday, 6=Sunday
                    ws.cell(row=row, column=col).value = "WO"
            
            # give some people WFH on Wednesdays (for testing "skip WFH" logic)
            if name in ["Jaya Shree", "Rahul Kumar"]:
                for day in range(1, days_in_month + 1):
                    d = date(year, month_num, day)
                    col = day + 2
                    if d.weekday() == 2:  # Wednesday
                        ws.cell(row=row, column=col).value = "WFH"
            
            # Add formulas for totals so they update automatically when opened in Excel
            last_col = openpyxl.utils.get_column_letter(days_in_month + 2)
            ws.cell(row=row, column=total_start).value = f'=COUNTIF(C{row}:{last_col}{row}, "CL")'
            ws.cell(row=row, column=total_start + 1).value = f'=COUNTIF(C{row}:{last_col}{row}, "SL")'
            ws.cell(row=row, column=total_start + 2).value = f'=COUNTIF(C{row}:{last_col}{row}, "PL")'
            ws.cell(row=row, column=total_start + 3).value = f'=COUNTIF(C{row}:{last_col}{row}, "EL")'
            ws.cell(row=row, column=total_start + 4).value = f'=COUNTIF(C{row}:{last_col}{row}, "HD")'
            ws.cell(row=row, column=total_start + 5).value = f'=COUNTIF(C{row}:{last_col}{row}, "WFH")'
            ws.cell(row=row, column=total_start + 6).value = f'=COUNTIF(C{row}:{last_col}{row}, "WO")'
            ws.cell(row=row, column=total_start + 7).value = f'=COUNTIF(C{row}:{last_col}{row}, "CO")'
            
            # Sum up CL, SL, PL, EL, and add half of HD
            cl_col = openpyxl.utils.get_column_letter(total_start)
            el_col = openpyxl.utils.get_column_letter(total_start + 3)
            hd_col = openpyxl.utils.get_column_letter(total_start + 4)
            ws.cell(row=row, column=total_start + 8).value = f'=SUM({cl_col}{row}:{el_col}{row}) + ({hd_col}{row}*0.5)'

        
        # mark a couple of holidays

        # June 15 = some office holiday, July 4 = another one (just for testing)
        if month_num == 6:
            holiday_day = 15
        else:
            holiday_day = 4
        
        for i in range(len(team)):
            row = i + 3
            ws.cell(row=row, column=holiday_day + 2).value = "OH"
    
    wb.save(path)
    print(f"Done! Saved to: {path}")
    print(f"  Sheets: {', '.join([s for s in wb.sheetnames])}")
    print(f"  Team members: {len(team)}")


def create_test_google_sheet(sheet_name):
    print(f"Initializing Google Sheet '{sheet_name}' with test data...")
    client = get_gspread_client()
        
    try:
        sh = client.open(sheet_name)
        print(f"Found existing Google Sheet '{sheet_name}'.")
    except gspread.SpreadsheetNotFound:
        # Create it!
        print(f"Google Sheet '{sheet_name}' not found. Creating a new one...")
        sh = client.create(sheet_name)
        print(f"Created new spreadsheet. Owner/Account: {client.auth.signer_email}")
        print("IMPORTANT: Share the sheet with this email as Editor so the script can access it!")
        
    team = [
        ("MDM-BAU", "Kavivanan"),
        ("MDM-BAU", "Leando Iruthayaraj"),
        ("MDM-BAU", "Vibra Narayanan"),
        ("MDM-BAU", "Aravind"),
        ("MDM-BAU", "Jaya Shree"),
        ("MDM - GROW", "Vignesh Pugalendhi"),
        ("MDM - GROW", "Sujitha"),
        ("MDM - GROW", "Balaji Loganathan"),
        ("MDM - GROW", "Ebron Stalin"),
        ("MDM - GROW", "Rahul Kumar"),
        ("MDM - GROW", "Harivarshan"),
        ("MDM - GROW", "BharaniDharan"),
        ("MDM - GROW", "Sandhiya"),
        ("MDM - GROW", "Jose"),
        ("CRM - Next Gen", "Dravid"),
        ("CRM - Next Gen", "Soma Sai Pavan"),
        ("CRM - Next Gen", "Rathimeena"),
        ("CRM - Next Gen", "Pavithra"),
        ("CRM - Next Gen", "Jefflina"),
        ("CRM - Next Gen", "Manoj"),
        ("CDS", "Arshad"),
        ("CDS", "Dushyanth"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "Hari Vembeina"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "JP"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "Jelsi"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "Mahesh Reddy"),
        ("PowerBI / Tableau / Alteryx / Devops Admin", "Sandeep Reddy"),
        ("", "Alafia Yusuf Fidvi"),
        ("", "Arun R"),
        ("", "Swaminathan"),
        ("", "Vasanth"),
    ]
    
    # Create sheets for Jun and Jul 2026
    for month_num in [6, 7]:
        year = 2026
        month_abbr = calendar.month_abbr[month_num]
        ws_name = f"{month_abbr} {year}"
        
        # Check if worksheet exists, if not create
        try:
            ws = sh.worksheet(ws_name)
            # clear it
            ws.clear()
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=ws_name, rows=100, cols=50)
            
        days_in_month = calendar.monthrange(year, month_num)[1]
        
        grid = []
        
        # Row 1
        row1 = ["Stream", "Team Member Name"]
        for day in range(1, days_in_month + 1):
            d = date(year, month_num, day)
            row1.append(d.strftime("%a"))
        row1.extend([
            "Total CL", "Total SL", "Total PL", "Total EL", "Total HD",
            "Total WFH", "Total Week Off", "Total Comp Off", "Total Leave Days"
        ])
        grid.append(row1)
        
        # Row 2
        row2 = ["", ""]
        for day in range(1, days_in_month + 1):
            row2.append(day)
        row2.extend([""] * 9)
        grid.append(row2)
        
        # Row 3+
        for i, (stream, name) in enumerate(team):
            row = [stream, name]
            for day in range(1, days_in_month + 1):
                d = date(year, month_num, day)
                val = ""
                if d.weekday() in [5, 6]:
                    val = "WO"
                elif name in ["Jaya Shree", "Rahul Kumar"] and d.weekday() == 2:
                    val = "WFH"
                elif day == 15 and month_num == 6:
                    val = "OH"
                elif day == 4 and month_num == 7:
                    val = "OH"
                row.append(val)
            
            row_num = i + 3
            last_col = openpyxl.utils.get_column_letter(days_in_month + 2)
            
            # Column letters for calculations
            total_start_col = days_in_month + 3
            cl_col_let = openpyxl.utils.get_column_letter(total_start_col)
            el_col_let = openpyxl.utils.get_column_letter(total_start_col + 3)
            hd_col_let = openpyxl.utils.get_column_letter(total_start_col + 4)
            
            row.extend([
                f'=COUNTIF(C{row_num}:{last_col}{row_num}, "CL")',
                f'=COUNTIF(C{row_num}:{last_col}{row_num}, "SL")',
                f'=COUNTIF(C{row_num}:{last_col}{row_num}, "PL")',
                f'=COUNTIF(C{row_num}:{last_col}{row_num}, "EL")',
                f'=COUNTIF(C{row_num}:{last_col}{row_num}, "HD")',
                f'=COUNTIF(C{row_num}:{last_col}{row_num}, "WFH")',
                f'=COUNTIF(C{row_num}:{last_col}{row_num}, "WO")',
                f'=COUNTIF(C{row_num}:{last_col}{row_num}, "CO")',
                f'=SUM({cl_col_let}{row_num}:{el_col_let}{row_num}) + ({hd_col_let}{row_num}*0.5)'
            ])
            grid.append(row)


            
        # Update sheet in one call
        ws.update(grid, f"A1:{gspread.utils.rowcol_to_a1(len(grid), len(grid[0]))}")
        print(f"  Initialized worksheet '{ws_name}'")
        
    print("Google Sheet initialization completed successfully!")


# ============================================================
# Run tests when this script is executed directly
# ============================================================

if __name__ == "__main__":
    
    print("=" * 60)
    print("  update_excel.py — Test Run")
    print("=" * 60)
    
    if USE_GOOGLE_SHEETS:
        print("\n>>> Running tests in GOOGLE SHEETS mode <<<")
        print(f"  Target Google Sheet: {GOOGLE_SHEET_NAME}")
        print(f"  Credentials file: {GOOGLE_CREDENTIALS_FILE}")
        
        try:
            create_test_google_sheet(GOOGLE_SHEET_NAME)
        except Exception as e:
            print(f"\nERROR: Could not initialize test Google Sheet: {e}")
            print("Please make sure you have created the service account, downloaded the JSON key file,")
            print("and shared the spreadsheet with the service account email.")
            import sys
            sys.exit(1)
            
        excel_path = None
    else:
        print("\n>>> Running tests in LOCAL EXCEL mode <<<")
        excel_path = os.path.join(script_dir, config["excel_path"])
        
        # --- create the test file ---
        print("\n[Setup] Creating test Excel file...")
        create_test_excel(excel_path)
        
    # --- Test 1: basic leave write ---
    print("\n" + "-" * 40)
    print("[Test 1] Write SL for 'Kavivanan' on June 16 (Tuesday)")
    success, msg = write_leave(excel_path, "Kavivanan", date(2026, 6, 16), "SL")
    print(f"  -> Result: success={success}, {msg}")
    
    # --- Test 2: try to write on a WFH day (should skip) ---
    print("\n" + "-" * 40)
    print("[Test 2] Try CL for 'Jaya Shree' on June 17 (Wednesday = her WFH day)")
    success, msg = write_leave(excel_path, "Jaya Shree", date(2026, 6, 17), "CL")
    print(f"  -> Result: success={success}, {msg}")
    
    # --- Test 3: fuzzy name matching ---
    print("\n" + "-" * 40)
    print("[Test 3] Fuzzy match: 'Rahul' should find 'Rahul Kumar'")
    success, msg = write_leave(excel_path, "Rahul", date(2026, 6, 22), "PL")
    print(f"  -> Result: success={success}, {msg}")
    
    # --- Test 4: don't overwrite existing leave ---
    print("\n" + "-" * 40)
    print("[Test 4] Try to overwrite — write CL on June 16 (already has SL from Test 1)")
    success, msg = write_leave(excel_path, "Kavivanan", date(2026, 6, 16), "CL")
    print(f"  -> Result: success={success}, {msg}")
    
    # --- Test 5: correction should overwrite ---
    print("\n" + "-" * 40)
    print("[Test 5] Correction — overwrite SL with PL on June 16 (is_correction=True)")
    success, msg = write_leave(excel_path, "Kavivanan", date(2026, 6, 16), "PL", is_correction=True)
    print(f"  -> Result: success={success}, {msg}")
    
    # --- Test 6: skip weekend (WO) ---
    print("\n" + "-" * 40)
    print("[Test 6] Try CL on Saturday June 20 (should be WO)")
    success, msg = write_leave(excel_path, "Sujitha", date(2026, 6, 20), "CL")
    print(f"  -> Result: success={success}, {msg}")
    
    # --- Test 7: skip holiday (OH) ---
    print("\n" + "-" * 40)
    print("[Test 7] Try SL on June 15 (office holiday = OH)")
    success, msg = write_leave(excel_path, "Dravid", date(2026, 6, 15), "SL")
    print(f"  -> Result: success={success}, {msg}")
    
    # --- Test 8: multi-day leave ---
    print("\n" + "-" * 40)
    print("[Test 8] Multi-day EL for 'Rathimeena' — June 22 to June 25 (Mon-Thu)")
    results = write_leave_multi_day(excel_path, "Rathimeena", date(2026, 6, 22), date(2026, 6, 25), "EL")
    print("\n  Summary:")
    for r in results:
        status = "OK" if r["success"] else r["message"]
        print(f"    {r['date'].strftime('%a %d %b')}: {status}")
    
    # --- Test 9: unknown name (should fail to match) ---
    print("\n" + "-" * 40)
    print("[Test 9] Unknown name: 'John Doe' (shouldn't match anyone)")
    success, msg = write_leave(excel_path, "John Doe", date(2026, 6, 18), "CL")
    print(f"  -> Result: success={success}, {msg}")
    
    # --- Test 10: another fuzzy match ---
    print("\n" + "-" * 40)
    print("[Test 10] Fuzzy match: 'S. Reddy' -> should find 'Sandeep Reddy'")
    success, msg = write_leave(excel_path, "S. Reddy", date(2026, 6, 18), "CL")
    print(f"  -> Result: success={success}, {msg}")
    
    print("\n" + "=" * 60)
    if USE_GOOGLE_SHEETS:
        print(f"  All tests done! Open Google Sheet '{GOOGLE_SHEET_NAME}' to verify.")
    else:
        print("  All tests done! Open test_leave_tracker.xlsx to verify.")
    print("=" * 60)
